import openpyxl

values=[]
workbook=openpyxl.load_workbook("C:\\Users\LENOVO\Downloads\excel2.xlsx")//location of  excel file
sh1=workbook['Sheet1']
//taking the ongoing sheet
row=sh1.max_row
//selecting max rows
column=sh1.max_column
//selecting max column


for i in range(1,row+1):
    moni=sh1.cell(i+1,3).value
    if sh1.cell(i + 1, 3).value is None:
        pass
    else:
        values.append(sh1.cell(i + 1, 3).value)
print(values)

ayush=[]
from textblob import TextBlob
for i in range(len(values)):
    ayush.append(TextBlob(values[i]).sentiment.polarity)
print(ayush)
verybad=0
bad=0
good=0
verygood=0
boss=[]
for i in range(len(ayush)):
    if ayush[i]<= -0.5:
        boss.append(values[i])
        verybad=verybad+1
    if 0 > ayush[i] >= -0.5:
        bad=bad+1
    if 0.5 > ayush[i] > 0:
        good = good+1
for i in range(len(ayush)):
    if ayush[i] >= 0.5 :
        verygood=verygood +1
print(verybad,bad,good,verygood)
print(boss)
import  matplotlib.pyplot as plt
x=[verybad,bad,good,verygood]
la=["verybad","bad","good","verygood"]
plt.pie(x,labels=la)
plt.show()





